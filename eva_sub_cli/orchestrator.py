#!/usr/bin/env python
import csv
import json
import os
from collections import defaultdict

import requests
from ebi_eva_common_pyutils.config import WritableConfig
from ebi_eva_common_pyutils.logger import logging_config
from openpyxl.reader.excel import load_workbook

from eva_sub_cli import SUB_CLI_CONFIG_FILE, __version__
from eva_sub_cli.exceptions.submission_not_found_exception import SubmissionNotFoundException
from eva_sub_cli.exceptions.submission_status_exception import SubmissionStatusException
from eva_sub_cli.submission_ws import SubmissionWSClient
from eva_sub_cli.submit import StudySubmitter, SUB_CLI_CONFIG_KEY_SUBMISSION_ID
from eva_sub_cli.validators.docker_validator import DockerValidator
from eva_sub_cli.validators.native_validator import NativeValidator
from eva_sub_cli.validators.validator import READY_FOR_SUBMISSION_TO_EVA

VALIDATE = 'validate'
SUBMIT = 'submit'
DOCKER = 'docker'
NATIVE = 'native'

logger = logging_config.get_logger(__name__)


def get_vcf_files(mapping_file):
    vcf_files = []
    with open(mapping_file) as open_file:
        reader = csv.DictReader(open_file, delimiter=',')
        for row in reader:
            vcf_files.append(row['vcf'])
    return vcf_files


def get_project_title_and_create_vcf_files_mapping(submission_dir, vcf_files, reference_fasta, metadata_json, metadata_xlsx):
    mapping_file = os.path.join(submission_dir, 'vcf_mapping_file.csv')
    with open(mapping_file, 'w') as open_file:
        writer = csv.writer(open_file, delimiter=',')
        writer.writerow(['vcf', 'fasta', 'report'])

        vcf_files_mapping = []
        if vcf_files and reference_fasta:
            for vcf_file in vcf_files:
                vcf_files_mapping.append([os.path.abspath(vcf_file), os.path.abspath(reference_fasta)])
            if metadata_json:
                project_title, _ = get_project_and_vcf_fasta_mapping_from_metadata_json(metadata_json, False)
            elif metadata_xlsx:
                project_title, _ = get_project_and_vcf_fasta_mapping_from_metadata_xlsx(metadata_xlsx, False)
        elif metadata_json:
            project_title, vcf_files_mapping = get_project_and_vcf_fasta_mapping_from_metadata_json(metadata_json, True)
        elif metadata_xlsx:
            project_title, vcf_files_mapping = get_project_and_vcf_fasta_mapping_from_metadata_xlsx(metadata_xlsx, True)

        for mapping in vcf_files_mapping:
            writer.writerow(mapping)

    return project_title, mapping_file


def get_project_and_vcf_fasta_mapping_from_metadata_json(metadata_json, mapping_req=False):
    with open(metadata_json) as file:
        json_metadata = json.load(file)

        if 'project' in json_metadata:
            if 'title' in json_metadata['project']:
                project_title = json_metadata['project']['title']

        vcf_fasta_report_mapping = []
        if mapping_req:
            analysis_alias_dict = defaultdict(dict)
            for analysis in json_metadata['analysis']:
                analysis_alias_dict[analysis['analysisAlias']]['referenceFasta'] = analysis['referenceFasta']
                analysis_alias_dict[analysis['analysisAlias']]['assemblyReport'] = analysis['assemblyReport'] \
                    if 'assemblyReport' in analysis else ''

            for file_dict in json_metadata['files']:
                reference_fasta = analysis_alias_dict[file_dict['analysisAlias']]['referenceFasta']
                assembly_report = analysis_alias_dict[file_dict['analysisAlias']]['assemblyReport']
                vcf_fasta_report_mapping.append([os.path.abspath(file_dict['fileName']),
                                                 os.path.abspath(reference_fasta),
                                                 os.path.abspath(assembly_report) if assembly_report else ''])

    return project_title, vcf_fasta_report_mapping


def get_project_and_vcf_fasta_mapping_from_metadata_xlsx(metadata_xlsx, mapping_req=False):
    workbook = load_workbook(metadata_xlsx)

    project_sheet = workbook['Project']
    project_headers = {}
    for cell in project_sheet[1]:
        project_headers[cell.value] = cell.column
    project_title = project_sheet.cell(row=2, column=project_headers['Project Title']).value

    vcf_fasta_report_mapping = []
    if mapping_req:
        analysis_alias_sheet = workbook['Analysis']
        analysis_headers = {}
        for cell in analysis_alias_sheet[1]:
            analysis_headers[cell.value] = cell.column - 1

        analysis_alias_dict = {}
        for row in analysis_alias_sheet.iter_rows(min_row=2, values_only=True):
            analysis_alias = row[analysis_headers['Analysis Alias']]
            reference_fasta = row[analysis_headers['Reference Fasta Path']]
            analysis_alias_dict[analysis_alias] = reference_fasta

        files_sheet = workbook['Files']
        files_headers = {}
        for cell in files_sheet[1]:
            files_headers[cell.value] = cell.column - 1

        for row in files_sheet.iter_rows(min_row=2, values_only=True):
            file_name = os.path.abspath(row[files_headers['File Name']])
            analysis_alias = row[files_headers['Analysis Alias']]
            reference_fasta = os.path.abspath(analysis_alias_dict[analysis_alias])
            if not (file_name and os.path.isfile(file_name)):
                raise FileNotFoundError(f'The variant file {file_name} provided in spreadsheet {metadata_xlsx} does not exist')
            if not (reference_fasta and os.path.isfile(reference_fasta)):
                raise FileNotFoundError(f'The reference fasta {reference_fasta} in spreadsheet {metadata_xlsx} does not exist')
            vcf_fasta_report_mapping.append([os.path.abspath(file_name), os.path.abspath(reference_fasta), ''])

    return project_title, vcf_fasta_report_mapping


def check_validation_required(tasks, sub_config, username=None, password=None):
    # Validation is mandatory so if submit is requested then VALIDATE must have run before or be requested as well
    if SUBMIT in tasks:
        if not sub_config.get(READY_FOR_SUBMISSION_TO_EVA, False):
            return True
        submission_id = sub_config.get(SUB_CLI_CONFIG_KEY_SUBMISSION_ID, None)
        if submission_id:
            try:
                submission_status = SubmissionWSClient(username, password).get_submission_status(submission_id)
                if submission_status == 'FAILED':
                    return True
                else:
                    return False
            except requests.HTTPError as ex:
                if ex.response.status_code == 404:
                    logger.error(
                        f'Submission with id {submission_id} could not be found: '
                        f'status code: {ex.response.status_code} response: {ex.response.text}')
                    raise SubmissionNotFoundException(f'Submission with id {submission_id} could not be found')
                else:
                    logger.error(f'Error occurred while getting status of the submission with Id {submission_id}: '
                                 f'status code: {ex.response.status_code} response: {ex.response.text}')
                    raise SubmissionStatusException(f'Error occurred while getting status of the submission '
                                                    f'with Id {submission_id}')

        logger.info(f'submission id not found in config. This might be the first time user is submitting')
        return False


def orchestrate_process(submission_dir, vcf_files, reference_fasta, metadata_json, metadata_xlsx,
                        tasks, executor, username=None, password=None, shallow_validation=False, **kwargs):
    # load config
    config_file_path = os.path.join(submission_dir, SUB_CLI_CONFIG_FILE)
    sub_config = WritableConfig(config_file_path, version=__version__)

    metadata_file = metadata_json or metadata_xlsx
    if not os.path.exists(os.path.abspath(metadata_file)):
        raise FileNotFoundError(f'The provided metadata file {os.path.abspath(metadata_file)} does not exist')

    if metadata_json:
        metadata_json = os.path.abspath(metadata_json)
    if metadata_xlsx:
        metadata_xlsx = os.path.abspath(metadata_xlsx)

    # Get the provided Project Title and VCF files mapping (VCF, Fasta and Report)
    project_title, vcf_files_mapping = get_project_title_and_create_vcf_files_mapping(
        submission_dir, vcf_files, reference_fasta, metadata_json, metadata_xlsx
    )
    vcf_files = get_vcf_files(vcf_files_mapping)

    if VALIDATE not in tasks and check_validation_required(tasks, sub_config, username, password):
        tasks.append(VALIDATE)

    if VALIDATE in tasks:
        if executor == DOCKER:
            validator = DockerValidator(vcf_files_mapping, submission_dir, project_title, metadata_json, metadata_xlsx,
                                        shallow_validation=shallow_validation, submission_config=sub_config)
        # default to native execution
        else:
            validator = NativeValidator(vcf_files_mapping, submission_dir, project_title, metadata_json, metadata_xlsx,
                                        shallow_validation=shallow_validation, submission_config=sub_config)
        with validator:
            validator.validate_and_report()
            if not metadata_json:
                metadata_json = os.path.join(validator.output_dir, 'metadata.json')
            sub_config.set('metadata_json', value=metadata_json)
            sub_config.set('vcf_files', value=vcf_files)

    if SUBMIT in tasks:
        with StudySubmitter(submission_dir, submission_config=sub_config, username=username, password=password) as submitter:
            submitter.submit()
