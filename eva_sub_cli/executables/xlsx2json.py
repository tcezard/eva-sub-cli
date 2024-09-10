import argparse
import datetime
import json
from json import JSONEncoder

import openpyxl
import yaml
from ebi_eva_common_pyutils.logger import logging_config
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

WORKSHEETS_KEY_NAME = 'worksheets'
REQUIRED_HEADERS_KEY_NAME = 'required'
OPTIONAL_HEADERS_KEY_NAME = 'optional'
HEADERS_KEY_ROW = 'header_row'
CAST_KEY_NAME = 'cast'
PROJECT = 'Project'
SAMPLE = 'Sample'
ANALYSIS_ALIAS_KEY = 'Analysis Alias'
SAMPLE_NAME_IN_VCF_KEY = 'Sample Name in VCF'
SAMPLE_ACCESSION_KEY = 'Sample Accession'
SAMPLE_NAME_KEY = 'BioSample Name'
SCIENTIFIC_NAME_KEY = 'Scientific Name'
SPECIES = 'species'

logger = logging_config.get_logger(__name__)


class DateTimeEncoder(JSONEncoder):
    # Override the default method
    def default(self, obj):
        if isinstance(obj, (datetime.date, datetime.datetime)):
            return obj.isoformat()


class XlsxParser:
    """
    Base parser for Excel file for the fields from worksheets defined in a configuration file.
    It implements the base functionality for opening and validating the spreadsheet.
    """

    def __init__(self, xlsx_filename, conf_filename):
        """
        Constructor
        :param xlsx_filename: Excel file path
        :type xlsx_filename: basestring
        :param conf_filename: configuration file path
        :type conf_filename: basestring
        """
        with open(conf_filename, 'r') as conf_file:
            self.xlsx_conf = yaml.safe_load(conf_file)
        try:
            self.workbook = load_workbook(xlsx_filename, read_only=True)
        except Exception as e:
            self.add_error(f'Error loading {xlsx_filename}: {repr(e)}')
            self.file_loaded = False
            return
        self.worksheets = []
        self._active_worksheet = None
        self.row_offset = {}
        self.headers = {}
        self.file_loaded = True
        self.errors = []
        self.valid_worksheets()

    @property
    def active_worksheet(self):
        return self._active_worksheet

    @active_worksheet.setter
    def active_worksheet(self, worksheet):
        if self.worksheets is None:
            self.valid_worksheets()
        if worksheet not in self.worksheets:
            self.add_error(f'Tried to access an invalid worksheet {worksheet}', sheet=worksheet)
            return

        self._active_worksheet = worksheet

    def valid_worksheets(self):
        """
        Get the list of the names of worksheets which have the expected title and header row.
        """
        sheet_titles = self.workbook.sheetnames

        for title in self.xlsx_conf[WORKSHEETS_KEY_NAME]:
            # Check worksheet exists
            if title not in sheet_titles:
                continue

            # Check number of rows
            worksheet = self.workbook[title]
            header_row = self.xlsx_conf[title].get(HEADERS_KEY_ROW, 1)
            if worksheet.max_row < header_row + 1:
                continue
            # Store headers and worksheet title
            self.headers[title] = [cell.value if cell.value is None else cell.value.strip()
                                   for cell in worksheet[header_row]]
            self.worksheets.append(title)

    @staticmethod
    def cast_value(value, type_name):
        # Do not cast None values
        if type_name and value is not None:
            if type_name == 'string':
                return str(value)
            if type_name == 'boolean':
                return str(value).lower() in ['true', '1', 't', 'y', 'yes']
            if type_name == 'list':
                # split and remove empty values
                return [XlsxParser.trim_value(v)
                        for v in value.split(',') if v]
            if type_name == 'date':
                return XlsxParser.serialize(value)
        return value

    @staticmethod
    def trim_value(value):
        """Remove whitespace from the start and end of cells"""
        if isinstance(value, str):
            return value.strip()
        return value

    def base_row_offset(self, worksheet):
        return self.xlsx_conf[worksheet].get(HEADERS_KEY_ROW, 1)

    def get_rows(self):
        """
        Retrieve all the data rows.
        :return: list of hash containing all the REQUIRED and OPTIONAL fields as keys
                and the corresponding data as values
        :rtype: list
        """
        worksheet = self.active_worksheet
        if worksheet is None:
            logger.warning('No worksheet is specified!')
            return None

        if worksheet not in self.row_offset:
            self.row_offset[worksheet] = self.base_row_offset(worksheet)
        self.row_offset[worksheet] += 1

        required_headers = self.xlsx_conf[worksheet].get(REQUIRED_HEADERS_KEY_NAME, {})
        optional_headers = self.xlsx_conf[worksheet].get(OPTIONAL_HEADERS_KEY_NAME, {})

        rows = []

        for row in self.workbook[worksheet].iter_rows(min_row=self.row_offset[worksheet]):
            num_cells = 0
            for cell in row:
                num_cells += 1

            data = {}
            has_notnull = False
            for header in list(required_headers.keys()) + list(optional_headers.keys()):
                header_index = num_cells
                if header in self.headers[worksheet]:
                    header_index = self.headers[worksheet].index(header)
                if header_index >= num_cells:
                    data[header] = None
                    continue

                cell = row[header_index]
                if cell.value is not None:
                    has_notnull = True
                data[header] = self.trim_value(self.cast_value(
                    cell.value, self.xlsx_conf[worksheet].get(CAST_KEY_NAME, {}).get(header)
                ))

            if has_notnull:
                data['row_num'] = self.row_offset[worksheet]
                rows.append(data)

            # no data on this row, continue to next
            self.row_offset[worksheet] += 1

        return rows

    def get_project_json_data(self):
        project_data = self.get_rows()
        if len(project_data) > 1:
            logger.warning(f"{PROJECT} worksheet expects a single row of info but more than one found in the file. "
                           f"Only the first row's data will be taken into consideration")

        first_row = project_data[0]
        first_row.pop('row_num')

        json_key = self.xlsx_conf[WORKSHEETS_KEY_NAME][PROJECT]
        json_value = {self.translate_header(PROJECT, k): v for k, v in first_row.items() if v is not None}
        return {json_key: json_value}

    def get_biosample_object(self, data):
        sample_name = self.xlsx_conf[SAMPLE][OPTIONAL_HEADERS_KEY_NAME][SAMPLE_NAME_KEY]
        scientific_name = self.xlsx_conf[SAMPLE][OPTIONAL_HEADERS_KEY_NAME][SCIENTIFIC_NAME_KEY]

        # BioSample expects any of organism or species field
        if scientific_name in data:
            data[SPECIES] = data[scientific_name]
        # BioSample name goes in its own attribute, not part of characteristics
        biosample_name = data.pop(sample_name, None)

        # For all characteristics, BioSample expects value in arrays of objects
        biosample_object = {
            'characteristics': {k: [{'text': self.serialize(v)}] for k, v in data.items()}
        }
        if biosample_name is not None:
            biosample_object['name'] = biosample_name

        return biosample_object

    @staticmethod
    def serialize(value):
        """Create a text representation of the value provided"""
        if isinstance(value, datetime.date):
            return value.strftime('%Y-%m-%d')
        return str(value)

    def get_sample_data_with_split_analysis_alias(self, data, analysis_alias, sample_name_in_vcf):
        analysis_alias_list = data[analysis_alias].split(',')
        sample_in_vcf_val = data[sample_name_in_vcf]
        return {analysis_alias: analysis_alias_list, sample_name_in_vcf: sample_in_vcf_val}

    def get_sample_json_data(self):
        json_key = self.xlsx_conf[WORKSHEETS_KEY_NAME][SAMPLE]
        sample_json = {json_key: []}
        for row in self.get_rows():
            row_num = row.pop('row_num')
            json_value = {self.translate_header(SAMPLE, k): v for k, v in row.items() if v is not None}
            bio_sample_acc = self.xlsx_conf[SAMPLE][OPTIONAL_HEADERS_KEY_NAME][SAMPLE_ACCESSION_KEY]

            analysis_alias = self.xlsx_conf[SAMPLE][REQUIRED_HEADERS_KEY_NAME][ANALYSIS_ALIAS_KEY]
            sample_name_in_vcf = self.xlsx_conf[SAMPLE][REQUIRED_HEADERS_KEY_NAME][SAMPLE_NAME_IN_VCF_KEY]
            sample_data = self.get_sample_data_with_split_analysis_alias(json_value, analysis_alias, sample_name_in_vcf)

            if bio_sample_acc in json_value and json_value[bio_sample_acc]:
                sample_data.update(bioSampleAccession=json_value[bio_sample_acc])
                sample_json[json_key].append(sample_data)
            else:
                json_value.pop(analysis_alias)
                json_value.pop(sample_name_in_vcf)

                biosample_obj = self.get_biosample_object(json_value)
                sample_data.update(bioSampleObject=biosample_obj)
                sample_json[json_key].append(sample_data)

        return sample_json

    def json(self, output_json_file):
        # If the file could not be loaded at all, return without generating JSON.
        if not self.file_loaded:
            return
        json_data = {}
        for title in self.xlsx_conf[WORKSHEETS_KEY_NAME]:
            self.active_worksheet = title
            if title == PROJECT:
                json_data.update(self.get_project_json_data())
            elif title == SAMPLE:
                sample_data = self.get_sample_json_data()
                json_data.update(sample_data)
            else:
                json_data[self.xlsx_conf[WORKSHEETS_KEY_NAME][title]] = []
                for row in self.get_rows():
                    # Remove the row number
                    row.pop('row_num')
                    # Remove any None and translate header name
                    json_data[self.xlsx_conf[WORKSHEETS_KEY_NAME][title]].append(
                        {self.translate_header(title, k): v for k, v in row.items() if v is not None}
                    )

        with open(output_json_file, 'w') as open_file:
            json.dump(json_data, open_file, cls=DateTimeEncoder)

    def translate_header(self, title, header):
        if header in self.xlsx_conf[title].get(REQUIRED_HEADERS_KEY_NAME, {}):
            return self.xlsx_conf[title][REQUIRED_HEADERS_KEY_NAME][header]
        if header in self.xlsx_conf[title].get(OPTIONAL_HEADERS_KEY_NAME, {}):
            return self.xlsx_conf[title][OPTIONAL_HEADERS_KEY_NAME][header]
        logger.warning(
            f'Header {header} in {title} sheet does not have translation in the config file. Leave it as is')
        return header

    def add_error(self, message, sheet='', row='', column=''):
        """Adds a conversion error using the same structure as other validation errors,
        and marks the spreadsheet as invalid."""
        self.errors.append({'sheet': sheet, 'row': row, 'column': column, 'description': message})

    def save_errors(self, errors_yaml_file):
        with open(errors_yaml_file, 'w') as open_file:
            yaml.safe_dump(self.errors, stream=open_file)


def create_xls_template_from_yaml(xlsx_filename, conf_filename):
    """
    Create a XLS empty template with the expected worksheet and header based on the configuration provided
    """
    with open(conf_filename, 'r') as conf_file:
        xlsx_conf = yaml.safe_load(conf_file)
    workbook = Workbook()
    workbook.remove(workbook.active)

    for title in xlsx_conf[WORKSHEETS_KEY_NAME]:
        # Create worksheet
        workbook.create_sheet(title)
        # For each Worksheet create the header row
        header_row = xlsx_conf[title].get(HEADERS_KEY_ROW, 1)
        header_col_index = 0
        for header_name in xlsx_conf[title].get(REQUIRED_HEADERS_KEY_NAME, []):
            header_col_index += 1
            cell = workbook[title].cell(column=header_col_index, row=header_row, value=header_name)
            cell.font = openpyxl.styles.Font(bold=True)
        for header_name in xlsx_conf[title].get(OPTIONAL_HEADERS_KEY_NAME, []):
            header_col_index += 1
            workbook[title].cell(column=header_col_index, row=header_row, value=header_name)
    workbook.save(xlsx_filename)


def main():
    arg_parser = argparse.ArgumentParser(
        description='Convert an xlsx spreadsheet containing expected sheets to json prior to validation')
    arg_parser.add_argument('--metadata_json', required=True, dest='metadata_json',
                            help='Path to output EVA metadata json file')
    arg_parser.add_argument('--metadata_xlsx', required=True, dest='metadata_xlsx', help='EVA metadata Excel file')
    arg_parser.add_argument('--conversion_configuration', dest='conversion_configuration',
                            help='Configuration file describing the expected content of the Excel spreadsheet')
    arg_parser.add_argument('--errors_yaml', required=True, dest='errors_yaml', help='Path to output errors as YAML')

    args = arg_parser.parse_args()
    logging_config.add_stdout_handler()
    parser = XlsxParser(args.metadata_xlsx, args.conversion_configuration)
    try:
        parser.json(args.metadata_json)
    except Exception as e:
        parser.add_error(f'An Error was raised while converting the spreadsheet to JSON: {repr(e)}')
    finally:
        parser.save_errors(args.errors_yaml)
